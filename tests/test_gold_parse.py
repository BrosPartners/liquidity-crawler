import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from tests.make_gold_fixture import build
from core.gold_parse import parse_gold_xlsx, BRAND_ORDER

FIX = os.path.join(os.path.dirname(__file__), "_gold_fixture.xlsx")
FIX_LAG = os.path.join(os.path.dirname(__file__), "_gold_fixture_lag.xlsx")


def _data():
    build(FIX)
    return parse_gold_xlsx(FIX)


def test_latest():
    d = _data()
    lt = d["latest"]
    assert lt["as_of"] == "2026-07-13"
    assert lt["sjc_sell"] == 148900000
    assert lt["world_gold_usd"] == 4076.4
    assert lt["usd_vnd"] == 26260
    assert lt["world_gold_vnd"] == 128000000
    # gap tự tính = 148900000 - 128000000
    assert lt["gap"] == 20900000
    assert abs(lt["pct_gap"] - (20900000 / 128000000)) < 1e-9


def test_brands():
    d = _data()
    comps = [b["company"] for b in d["latest"]["brands"]]
    # đủ 6 hãng, đúng thứ tự BRAND_ORDER
    assert comps == [c for c in BRAND_ORDER if c in comps]
    assert len(comps) == 6
    sjc = next(b for b in d["latest"]["brands"] if b["company"] == "SJC")
    assert sjc["sell"] == 148900000
    # PNJ phải lấy sjc_at_pnj_hn (148.9M) KHÔNG phải vang_15k (94.9M)
    pnj = next(b for b in d["latest"]["brands"] if b["company"] == "PNJ")
    assert pnj["sell"] == 148900000


def test_history_skips_junk():
    d = _data()
    # hàng rác cuối (date=None) không lọt vào history
    assert all(r["date"] for r in d["history"])
    # có ngày 2026-07-13
    assert any(r["date"] == "2026-07-13" for r in d["history"])


def _build_lag_fixture(path):
    """Dựng fixture riêng: world_gold_vnd trễ hơn sjc_sell/usd_vnd vài ngày.

    2026-07-10: đủ dữ liệu (world_gold_vnd + sjc_sell) -> gap tính được.
    2026-07-13: có sjc_sell + usd_vnd + world_gold_usd nhưng KHÔNG có
    world_gold_vnd (không có row trong sheet "Gia TG (VND-luong)").
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Gia TG (USD_oz)")
    ws.append(["Nguồn: ..."])
    ws.append(["date", "open_usd", "high_usd", "low_usd", "close_usd", "source"])
    ws.append(["2026-07-10", 4122.3, 4125.8, 4090.6, 4104.1, "yf"])
    ws.append(["2026-07-13", 4106.6, 4111.6, 4069.4, 4076.4, "yf"])

    # Chỉ có 1 hàng (07-10) -> 07-13 KHÔNG có world_gold_vnd (world gold trễ)
    ws = wb.create_sheet("Gia TG (VND-luong)")
    ws.append(["Nguồn: ..."])
    ws.append(["date", "close_usd", "usd_vnd", "close_vnd", "Gia SJC", "Gap giá vàng", "% gap giá vàng"])
    ws.append(["2026-07-10", 4104.1, 26290, 130000000, "#N/A", "#N/A", "#N/A"])

    ws = wb.create_sheet("Ty gia USD-VND")
    ws.append(["Nguồn: ..."])
    ws.append(["date", "usd_vnd", "source"])
    ws.append(["2026-07-10", 26290, "yf"])
    ws.append(["2026-07-13", 26260, "yf"])

    ws = wb.create_sheet("SJC")
    ws.append(["Nguồn: ..."])
    ws.append(["date", "sjc_mieng_buy_price", "sjc_mieng_1c_buy_price", "sjc_mieng_1l_buy_price",
               "sjc_mieng_5c_buy_price", "sjc_nhan_1c_buy_price", "sjc_nutrang_75_buy_price",
               "sjc_nutrang_99_buy_price", "sjc_nutrang_9999_buy_price", "sjc_mieng_sell_price"])
    ws.append(["2026-07-10", 146900000, None, None, None, None, None, None, None, 149900000])
    ws.append(["2026-07-13", 145900000, None, None, None, None, None, None, None, 148900000])

    ws = wb.create_sheet("Gia VN (tat ca)")
    ws.append(["Nguồn: ..."])
    ws.append(["id", "date", "company", "gold_type", "purity", "buy_price", "sell_price", "unit", "source"])
    ws.append([1, "2026-07-13", "SJC", "sjc_mieng", 999.9, 145900000, 148900000, "luong", "24h"])

    wb.save(path)


def test_latest_is_coherent_when_world_lags():
    _build_lag_fixture(FIX_LAG)
    d = parse_gold_xlsx(FIX_LAG)
    lt = d["latest"]
    # Ngày 07-13 KHÔNG có world_gold_vnd -> snapshot phải lùi về 07-10
    # (ngày cuối cùng có gap != None), KHÔNG lấy sjc_sell 07-13 trộn với
    # world_gold_vnd 07-10 (đó là bug cũ).
    assert lt["as_of"] == "2026-07-10"
    assert lt["sjc_sell"] == 149900000
    assert lt["world_gold_vnd"] == 130000000
    assert lt["gap"] == lt["sjc_sell"] - lt["world_gold_vnd"]
    assert abs(lt["pct_gap"] - (lt["gap"] / lt["world_gold_vnd"])) < 1e-9


if __name__ == "__main__":
    test_latest()
    test_brands()
    test_history_skips_junk()
    test_latest_is_coherent_when_world_lags()
    print("Tất cả test PASS")
