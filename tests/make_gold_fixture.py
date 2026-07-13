"""Dựng file xlsx nhỏ mô phỏng đúng cấu trúc gold_prices_all.xlsx để test parser."""
import openpyxl


def build(path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Gia TG (USD_oz)")
    ws.append(["Nguồn: ..."])
    ws.append(["date", "open_usd", "high_usd", "low_usd", "close_usd", "source"])
    ws.append(["2026-07-10", 4122.3, 4125.8, 4090.6, 4104.1, "yf"])
    ws.append(["2026-07-13", 4106.6, 4111.6, 4069.4, 4076.4, "yf"])

    ws = wb.create_sheet("Gia TG (VND-luong)")
    ws.append(["Nguồn: ..."])
    ws.append(["date", "close_usd", "usd_vnd", "close_vnd", "Gia SJC", "Gap giá vàng", "% gap giá vàng"])
    ws.append(["2026-07-13", 4076.4, 26260, 128000000, "#N/A", "#N/A", "#N/A"])
    ws.append([None, None, None, None, "#N/A", "#N/A", "#N/A"])  # hàng rác cuối

    ws = wb.create_sheet("Ty gia USD-VND")
    ws.append(["Nguồn: ..."])
    ws.append(["date", "usd_vnd", "source"])
    ws.append(["2026-07-10", 26290, "yf"])
    ws.append(["2026-07-13", 26260, "yf"])

    ws = wb.create_sheet("SJC")
    ws.append(["Nguồn: ..."])
    ws.append(["date", "sjc_mieng_buy_price", "sjc_mieng_1c_buy_price", "sjc_mieng_1l_buy_price",
               "sjc_mieng_5c_buy_price", "sjc_nhan_1c_buy_price", "sjc_nutrang_75_buy_price",
               "sjc_nutrang_99_buy_price", "sjc_nutrang_9999_buy_price", "sjc_mieng_sell_price"])
    ws.append(["2026-07-13", 145900000, None, None, None, None, None, None, None, 148900000])

    ws = wb.create_sheet("Gia VN (tat ca)")
    ws.append(["Nguồn: ..."])
    ws.append(["id", "date", "company", "gold_type", "purity", "buy_price", "sell_price", "unit", "source"])
    rows = [
        (1, "2026-07-13", "SJC", "sjc_mieng", 999.9, 145900000, 148900000, "luong", "24h"),
        (2, "2026-07-13", "SJC", "sjc_nhan_1c", 999.9, 145400000, 148400000, "luong", "24h"),
        (3, "2026-07-13", "DOJI", "doji_hn", 999.9, 145900000, 148900000, "luong", "24h"),
        (4, "2026-07-13", "BTMC", "btmc_sjc", 999.9, 145000000, 149900000, "luong", "24h"),
        (5, "2026-07-13", "BTMH", "btmh_mieng", 999.9, 144000000, 148000000, "luong", "24h"),
        (6, "2026-07-13", "PHUQUY", "phuquy_sjc", 999.9, 145500000, 148900000, "luong", "24h"),
        (7, "2026-07-13", "PNJ", "sjc_at_pnj_hn", 999.9, 145900000, 148900000, "luong", "24h"),
        (8, "2026-07-13", "PNJ", "vang_15k", 999.9, 85000000, 94900000, "luong", "24h"),
    ]
    for r in rows:
        ws.append(list(r))

    wb.save(path)


if __name__ == "__main__":
    build("tests/_gold_fixture.xlsx")
    print("OK fixture -> tests/_gold_fixture.xlsx")
