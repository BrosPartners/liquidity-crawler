"""Parse gold_prices_all.xlsx (do bot Telegram tạo) -> dữ liệu cho dashboard.

Đọc theo TÊN sheet/cột (không theo index). Ô '#N/A'/trống -> None.
gap / pct_gap TỰ TÍNH (sjc_sell - world_gold_vnd) để né '#N/A' trong file.
"""
from __future__ import annotations

import openpyxl

# Mỗi hãng chọn 1 loại vàng "miếng SJC-tương đương" để so sánh công bằng.
BRAND_GOLD_TYPE = {
    "SJC": "sjc_mieng",
    "DOJI": "doji_hn",
    "BTMC": "btmc_sjc",
    "BTMH": "btmh_mieng",
    "PHUQUY": "phuquy_sjc",
    "PNJ": "sjc_at_pnj_hn",
}
BRAND_ORDER = ["SJC", "DOJI", "PNJ", "BTMC", "BTMH", "PHUQUY"]


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _date_str(v):
    if v is None:
        return None
    try:
        return v.strftime("%Y-%m-%d")
    except AttributeError:
        s = str(v).strip()
        return s[:10] if s else None


def _sheet_map(ws, val_cols):
    """Trả {date_str: {col: number}} đọc theo tên cột, header ở hàng 2."""
    it = ws.iter_rows(min_row=2, values_only=True)
    try:
        header = list(next(it))
    except StopIteration:
        return {}
    if "date" not in header:
        return {}
    di = header.index("date")
    idx = {c: header.index(c) for c in val_cols if c in header}
    out = {}
    for r in it:
        d = _date_str(r[di])
        if not d:
            continue
        out[d] = {c: _num(r[i]) for c, i in idx.items()}
    return out


def parse_gold_xlsx(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    required = ["Gia TG (USD_oz)", "Gia TG (VND-luong)", "Ty gia USD-VND", "SJC", "Gia VN (tat ca)"]
    missing = [s for s in required if s not in wb.sheetnames]
    if missing:
        raise ValueError(f"gold xlsx thiếu sheet: {missing}. Sheet hiện có: {wb.sheetnames}")

    usd = _sheet_map(wb["Gia TG (USD_oz)"], ["close_usd"])
    vnd = _sheet_map(wb["Gia TG (VND-luong)"], ["close_vnd"])
    # File có thể dùng cột 'usd_vnd' (cũ) hoặc 4 cột nguồn mới (VCB/tự do/SBV/yfinance).
    fx = _sheet_map(wb["Ty gia USD-VND"], [
        "usd_vnd", "usd_vnd_VCB_fiinpro", "usd_vnd_tudo_fiinpro",
        "usd_vnd_SBV_fiinpro", "usd_vnd_yfinance",
    ])
    sjc = _sheet_map(wb["SJC"], ["sjc_mieng_sell_price"])

    all_dates = sorted(set(usd) | set(vnd) | set(fx) | set(sjc))
    history = []
    for d in all_dates:
        wg_usd = usd.get(d, {}).get("close_usd")
        wg_vnd = vnd.get(d, {}).get("close_vnd")
        sjc_sell = sjc.get(d, {}).get("sjc_mieng_sell_price")
        fxrow = fx.get(d, {})
        # Ưu tiên: usd_vnd (cũ) -> VCB -> tự do -> SBV -> yfinance.
        rate = (fxrow.get("usd_vnd") or fxrow.get("usd_vnd_VCB_fiinpro")
                or fxrow.get("usd_vnd_tudo_fiinpro") or fxrow.get("usd_vnd_SBV_fiinpro")
                or fxrow.get("usd_vnd_yfinance"))
        gap = pct = None
        if wg_vnd and sjc_sell:
            gap = sjc_sell - wg_vnd
            pct = gap / wg_vnd
        history.append({
            "date": d, "world_gold_usd": wg_usd, "world_gold_vnd": wg_vnd,
            "sjc_sell": sjc_sell, "usd_vnd": rate, "gap": gap, "pct_gap": pct,
        })

    ws = wb["Gia VN (tat ca)"]
    it = ws.iter_rows(min_row=2, values_only=True)
    h = list(next(it))
    ci = {n: h.index(n) for n in ("date", "company", "gold_type", "buy_price", "sell_price")}
    brands = []
    for r in it:
        comp = r[ci["company"]]
        if comp not in BRAND_GOLD_TYPE:
            continue
        if r[ci["gold_type"]] != BRAND_GOLD_TYPE[comp]:
            continue
        d = _date_str(r[ci["date"]])
        if not d:
            continue
        brands.append({"date": d, "company": comp,
                       "buy": _num(r[ci["buy_price"]]), "sell": _num(r[ci["sell_price"]])})

    return {"latest": _build_latest(history, brands), "history": history, "brands": brands}


def _build_latest(history, brands):
    # Snapshot lấy từ ngày mới nhất có ĐỦ dữ liệu so sánh (gap != None) để các
    # KPI nhất quán (gap == sjc_sell - world_gold_vnd). world_gold_vnd thường
    # trễ hơn sjc_sell vài ngày nên KHÔNG back-fill từng field độc lập.
    row = {}
    for r in reversed(history):
        if r.get("gap") is not None:
            row = r
            break
    if not row and history:
        row = history[-1]
    bmax = {}
    for b in brands:
        cur = bmax.get(b["company"])
        if cur is None or b["date"] > cur["date"]:
            bmax[b["company"]] = b
    brand_list = [{"company": c, "buy": bmax[c]["buy"], "sell": bmax[c]["sell"]}
                  for c in BRAND_ORDER if c in bmax]
    brands_as_of = max((bmax[c]["date"] for c in bmax), default=None)
    return {
        "as_of": row.get("date"),
        "world_gold_usd": row.get("world_gold_usd"),
        "world_gold_vnd": row.get("world_gold_vnd"),
        "sjc_sell": row.get("sjc_sell"),
        "gap": row.get("gap"),
        "pct_gap": row.get("pct_gap"),
        "usd_vnd": row.get("usd_vnd"),
        "brands": brand_list,
        "brands_as_of": brands_as_of,
    }
